# Copyright (C) 2026 gbao86 <tiktokthu10@gmail.com>
# This file is part of the chims project.
# Licensed under the GNU General Public License v3.0; see LICENSE for details.
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
import os
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from app.database import get_db
from app.auth.dependencies import get_current_user

FONT_FAMILY = "NotoSans"
FONT_BOLD = "NotoSans-Bold"

# Thư mục font được bundle cùng project — không phụ thuộc hệ điều hành
_FONTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "fonts")


def _register_pdf_fonts():
    if FONT_FAMILY in pdfmetrics.getRegisteredFontNames():
        return

    regular_path = os.path.join(_FONTS_DIR, "NotoSans-Regular.ttf")
    bold_path = os.path.join(_FONTS_DIR, "NotoSans-Bold.ttf")

    if not os.path.exists(regular_path):
        raise FileNotFoundError(
            f"Bundled font not found: {regular_path}\n"
            "Please ensure NotoSans-Regular.ttf and NotoSans-Bold.ttf "
            "exist in backend/app/assets/fonts/"
        )

    pdfmetrics.registerFont(TTFont(FONT_FAMILY, regular_path))
    pdfmetrics.registerFont(TTFont(FONT_BOLD, bold_path if os.path.exists(bold_path) else regular_path))

router = APIRouter()
COMPANY_NAME = "CÔNG TY TNHH THƯƠNG MẠI GBAO"
COMPANY_ADDRESS = "Địa chỉ: Phường Tân Hòa, TPHCM"
COMPANY_HOTLINE = "Chưa cập nhật"
COMPANY_EMAIL = "gbao74452@gmail.com"
COMPANY_WEBSITE = "CHIMS"


async def _build_summary(db):
    now = datetime.now(timezone.utc)
    start_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_week = now - timedelta(days=7)

    total_inventory = await db.inventory.count_documents({})
    low_stock = await db.inventory.count_documents({"status": {"$in": ["low_stock", "out_of_stock"]}})
    total_customers = await db.customers.count_documents({})
    total_suppliers = await db.suppliers.count_documents({})
    total_sales_orders = await db.sales_orders.count_documents({})
    total_purchase_orders = await db.purchase_orders.count_documents({})
    total_warranties = await db.warranties.count_documents({})

    sales_result = await db.sales_orders.aggregate([
        {"$match": {"created_at": {"$gte": start_month}}},
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}}},
    ]).to_list(1)
    sales_this_month = sales_result[0]["total"] if sales_result else 0

    po_result = await db.purchase_orders.aggregate([
        {"$match": {"created_at": {"$gte": start_month}, "status": "received"}},
        {"$group": {"_id": None, "total": {"$sum": "$total_amount"}}},
    ]).to_list(1)
    purchases_this_month = po_result[0]["total"] if po_result else 0

    activity_cursor = db.tickets.aggregate([
        {"$match": {"updated_at": {"$gte": start_week}}},
        {"$group": {"_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$updated_at"}}, "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
    ])
    weekly_activity = []
    async for doc in activity_cursor:
        weekly_activity.append({"date": doc["_id"], "count": doc["count"]})

    return {
        "total_inventory": total_inventory,
        "low_stock": low_stock,
        "total_customers": total_customers,
        "total_suppliers": total_suppliers,
        "total_sales_orders": total_sales_orders,
        "total_purchase_orders": total_purchase_orders,
        "total_warranties": total_warranties,
        "sales_this_month": sales_this_month,
        "purchases_this_month": purchases_this_month,
        "weekly_activity": weekly_activity,
    }


@router.get("/summary")
async def reports_summary(current_user: dict = Depends(get_current_user)):
    db = get_db()
    return await _build_summary(db)


def _style_excel_header(ws, end_col: int):
    header_fill = PatternFill("solid", fgColor="6366F1")
    white_font = Font(color="FFFFFF", bold=True)
    for col in range(1, end_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 24


@router.get("/summary.xlsx")
async def export_report_xlsx(current_user: dict = Depends(get_current_user)):
    db = get_db()
    summary = await _build_summary(db)
    _now = datetime.now(timezone(timedelta(hours=7)))
    _report_time = _now.strftime("%H:%M - %d/%m/%Y")
    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws = wb.create_sheet("Summary")
    ws.append([COMPANY_NAME])
    ws.append([COMPANY_ADDRESS])
    ws.append([f"Hotline: {COMPANY_HOTLINE}"])
    ws.append([f"Email: {COMPANY_EMAIL}"])
    ws.append([f"Website: {COMPANY_WEBSITE}"])
    ws.append([f"Th\u1eddi gian l\u1eadp b\u00e1o c\u00e1o: {_report_time}"])
    ws.append([""])
    ws.append(["Metric", "Value"])
    summary_rows = [
        ["Total Inventory", summary["total_inventory"]],
        ["Low Stock", summary["low_stock"]],
        ["Total Customers", summary["total_customers"]],
        ["Total Suppliers", summary["total_suppliers"]],
        ["Total Sales Orders", summary["total_sales_orders"]],
        ["Total Purchase Orders", summary["total_purchase_orders"]],
        ["Total Warranties", summary["total_warranties"]],
        ["Sales This Month", summary["sales_this_month"]],
        ["Purchases This Month", summary["purchases_this_month"]],
    ]
    for row in summary_rows:
        ws.append(row)
    _style_excel_header(ws, 2)
    for cell in ws[1] + ws[2] + ws[3] + ws[4] + ws[5] + ws[6]:
        cell.font = Font(bold=True)
    ws[8][0].font = Font(bold=True)
    ws[8][1].font = Font(bold=True)

    # Activity sheet
    act = wb.create_sheet("Weekly Activity")
    act.append(["Date", "Count"])
    for item in summary["weekly_activity"]:
        act.append([item["date"], item["count"]])
    _style_excel_header(act, 2)

    # Inventory sheet
    inv = wb.create_sheet("Inventory Detail")
    inv.append(["SKU", "Name", "Category", "Brand", "Stock", "Cost Price", "Unit Price", "Status", "Location"])
    async for item in db.inventory.find({}).sort("created_at", -1):
        inv.append([
            item.get("sku_code", ""), item.get("name", ""), item.get("category", ""), item.get("brand", ""),
            item.get("stock_quantity", 0), item.get("cost_price", 0), item.get("unit_price", 0), item.get("status", ""), item.get("location", "")
        ])
    _style_excel_header(inv, 9)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    _xlsx_name = _now.strftime("report_summary_%H-%M_%d%m%Y.xlsx")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{_xlsx_name}"'},
    )


@router.get("/summary.pdf")
async def export_report_pdf(current_user: dict = Depends(get_current_user)):
    _register_pdf_fonts()
    db = get_db()
    summary = await _build_summary(db)
    _now = datetime.now(timezone(timedelta(hours=7)))
    _report_time = _now.strftime("%H:%M - %d/%m/%Y")
    buffer = BytesIO()
    # landscape(A4) = 841.9 x 595.3 pt; usable width = 841.9 - 36 margins = ~806 pt
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=18, leftMargin=18,
        topMargin=18, bottomMargin=18,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleArial", parent=styles["Title"],
        fontName=FONT_BOLD, fontSize=18, leading=22,
    )
    heading_style = ParagraphStyle(
        "Heading2Arial", parent=styles["Heading2"],
        fontName=FONT_BOLD, fontSize=12, leading=16,
    )
    body_style = ParagraphStyle(
        "BodyArial", parent=styles["BodyText"],
        fontName=FONT_FAMILY, fontSize=10, leading=13,
    )
    # ── Truncate helper ───────────────────────────────────────────────────────
    # Cắt ngắn text theo số ký tự tối đa, thêm "…" nếu bị cắt.
    # Dùng plain string (không phải Paragraph) → chiều cao hàng cố định, đẹp khi in.
    def _t(text, max_chars: int) -> str:
        s = str(text) if text is not None else ""
        return s if len(s) <= max_chars else s[: max_chars - 1] + "…"

    # Giới hạn ký tự mỗi cột inventory (font 7pt, ~5.5pt/char)
    # SKU=10, Name=28, Cat=13, Brand=13, Stock=6, Cost=10, Unit=10, Status=11, Loc=14
    INV_MAX = [10, 28, 13, 13, 6, 10, 10, 11, 14]

    # Helper tạo TableStyle dùng chung
    def _tbl_style(header_color: str, alt_color: str) -> TableStyle:
        return TableStyle([
            ('FONTNAME',      (0, 0), (-1, -1), FONT_FAMILY),
            ('FONTNAME',      (0, 0), (-1,  0), FONT_BOLD),
            ('FONTSIZE',      (0, 0), (-1, -1), 7),
            ('LEADING',       (0, 0), (-1, -1), 9),
            ('BACKGROUND',    (0, 0), (-1,  0), colors.HexColor(header_color)),
            ('TEXTCOLOR',     (0, 0), (-1,  0), colors.white),
            ('GRID',          (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 5),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor(alt_color)]),
        ])

    ROW_H = 14  # chiều cao hàng cố định (pt) — tất cả hàng đồng đều

    timestamp_style = ParagraphStyle(
        "TimestampArial", fontName=FONT_BOLD, fontSize=9, leading=12,
        textColor=colors.HexColor('#6366f1'),
    )

    elements = [Paragraph("CHIMS Reports Summary", title_style), Spacer(1, 6)]
    elements += [
        Paragraph(COMPANY_NAME, heading_style),
        Paragraph(COMPANY_ADDRESS, body_style),
        Paragraph(f"Hotline: {COMPANY_HOTLINE}", body_style),
        Paragraph(f"Email: {COMPANY_EMAIL}", body_style),
        Paragraph(f"Website: {COMPANY_WEBSITE}", body_style),
        Paragraph(f"Th\u1eddi gian l\u1eadp b\u00e1o c\u00e1o: {_report_time}", timestamp_style),
        Spacer(1, 12),
    ]

    # ── Summary table ─────────────────────────────────────────────────────────
    summary_data = [
        ["Metric",                "Value"],
        ["Total Inventory",       str(summary["total_inventory"])],
        ["Low Stock",             str(summary["low_stock"])],
        ["Total Customers",       str(summary["total_customers"])],
        ["Total Suppliers",       str(summary["total_suppliers"])],
        ["Total Sales Orders",    str(summary["total_sales_orders"])],
        ["Total Purchase Orders", str(summary["total_purchase_orders"])],
        ["Total Warranties",      str(summary["total_warranties"])],
        ["Sales This Month",      str(summary["sales_this_month"])],
        ["Purchases This Month",  str(summary["purchases_this_month"])],
    ]
    summary_table = Table(
        summary_data,
        colWidths=[200, 120],
        rowHeights=[ROW_H] * len(summary_data),
        repeatRows=1, hAlign="LEFT",
    )
    summary_table.setStyle(_tbl_style('#6366f1', '#f3f4f6'))
    elements.append(summary_table)

    # ── Weekly activity table ─────────────────────────────────────────────────
    if summary["weekly_activity"]:
        elements += [Spacer(1, 12), Paragraph("Weekly Activity", heading_style)]
        act_data = [["Date", "Count"]] + [
            [x["date"], str(x["count"])] for x in summary["weekly_activity"]
        ]
        act_table = Table(
            act_data,
            colWidths=[160, 80],
            rowHeights=[ROW_H] * len(act_data),
            repeatRows=1, hAlign="LEFT",
        )
        act_table.setStyle(_tbl_style('#0f172a', '#f3f4f6'))
        elements.append(act_table)

    # ── Inventory detail table ────────────────────────────────────────────────
    # 9 cột, tổng ~745 pt (< 806 pt usable landscape)
    # SKU=65, Name=175, Cat=80, Brand=80, Stock=45, Cost=70, Unit=70, Status=70, Loc=90
    inv_col_widths = [65, 175, 80, 80, 45, 70, 70, 70, 90]
    elements += [Spacer(1, 12), Paragraph("Inventory Detail", heading_style)]

    inv_headers = ["SKU", "Name", "Category", "Brand", "Stock",
                   "Cost Price", "Unit Price", "Status", "Location"]
    inv_data = [inv_headers]
    async for item in db.inventory.find({}).sort("created_at", -1):
        row_raw = [
            item.get("sku_code", ""),
            item.get("name", ""),
            item.get("category", ""),
            item.get("brand", ""),
            str(item.get("stock_quantity", 0)),
            str(item.get("cost_price", 0)),
            str(item.get("unit_price", 0)),
            item.get("status", ""),
            item.get("location", ""),
        ]
        inv_data.append([_t(v, m) for v, m in zip(row_raw, INV_MAX)])

    inv_table = Table(
        inv_data,
        colWidths=inv_col_widths,
        rowHeights=[ROW_H] * len(inv_data),
        repeatRows=1,
        hAlign="LEFT",
    )
    inv_table.setStyle(_tbl_style('#111827', '#f9fafb'))
    elements.append(inv_table)

    doc.build(elements)
    buffer.seek(0)
    _now = datetime.now(timezone(timedelta(hours=7)))
    _pdf_name = _now.strftime("report_summary_%H-%M_%d%m%Y.pdf")
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{_pdf_name}"'},
    )

